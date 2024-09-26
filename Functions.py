import re
import tkinter
from datetime import datetime
import vk_api
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from tkinter import messagebox
from settings import *

from openpyxl.utils import exceptions

# from settings import *


def find_date_re(date):
    if isinstance(date, datetime):
        return date.strftime("%d.%m")
    dm = re.findall(r'[0-3]*\d\.[0-1]\d*', date)
    if dm:
        dm = dm[0]
        if len(dm) - dm.find(".") == 2:
            dm = dm[:dm.find(".") + 1] + "0" + dm[len(dm) - 1]
        if dm.find(".") == 1:
            dm = "0" + dm
    return dm


def getting_history(stopper, dict_of_dancers, session, chat_id, log_messages=None, infi=False):  # Возвращает лист с финальными списками
    offset = 0
    list_of_final_lists = []
    stopper_flag = True
    amount = len(dict_of_dancers) + 1
    list_of_dancers_array = dict_of_dancers.keys()
    while stopper_flag:
        chat_history = session.method("messages.getHistory", {"peer_id": chat_id, "count": 200, "offset": offset})
        offset += 200
        # chat_history это словарь с огромный количеством данных.
        # в нём есть лист chat_history['items'], Где хрянятся сообщения
        for message in chat_history['items']:
            if (len(message['attachments']) and message['attachments'][0]['type'] == 'poll' and find_date_re(
                    message['attachments'][0]['poll']['question'])):
                # ----------------------------------------------------------------------------------------------------------------------------
                # Три условия проверки на правильное голосование.
                # Первое условие сделано для оптимизации, его смысл таков:
                # Если в сообщении не было никакого вложения "== 0",
                # то условие не срабатывает сразу и сообщение пропускается
                # Второе условие просто проверяет, что вложение это голосование
                # Третье условие находит из названия опроса реальную дату
                # ----------------------------------------------------------------------------------------------------------------------------
                the_poll = message['attachments'][0]['poll']
                date = find_date_re(the_poll['question'])
                if (stopper == date):
                    stopper_flag = False
                    log_messages.insert(tkinter.END, f"Stopper is False. stopper = {stopper}")
                    log_messages.insert(tkinter.END, f"the_poll['question'] = {the_poll['question']}")
                    print("the_poll['question'] = ", the_poll['question'], "Stopper is False")
                    break
                else:
                    final_list = ["и" for i in range(amount)]
                    poll_id = the_poll['id']  # ID опроса
                    budu_id = the_poll['answers'][0]['id']  # ID ответов "Буду" в опросе
                    ne_budu_id = the_poll['answers'][1]['id']  # ID ответов "Не Буду" в опросе
                    poll_budu = session.method("polls.getVoters", {'poll_id': poll_id, "answer_ids": budu_id})
                    poll_ne_budu = session.method("polls.getVoters", {'poll_id': poll_id, "answer_ids": ne_budu_id})
                    budu_users = session.method("users.get",
                                                {"user_ids": ",".join(map(str, poll_budu[0]['users']['items']))})
                    ne_budu_users = session.method("users.get",
                                                   {"user_ids": ",".join(map(str, poll_ne_budu[0]['users']['items']))})
                    # all_the_users - это список всех словарей, где каждый элемент списка это человек
                    all_users = []
                    for i in range(len(budu_users)):
                        full_name = budu_users[i]['first_name'] + " " + budu_users[i]['last_name']
                        if full_name in list_of_dancers_array:
                            all_users.append(full_name)
                            # full_name - это имя человека
                            final_list[dict_of_dancers[full_name]] = 1
                            # В финальном списке каждый проголосовавший человек помечается единицей
                    for i in range(len(ne_budu_users)):
                        full_name = ne_budu_users[i]['first_name'] + " " + ne_budu_users[i]['last_name']
                        if full_name in list_of_dancers_array:
                            all_users.append(full_name)
                            final_list[dict_of_dancers[full_name]] = 'н'
                    final_list.insert(0, date)  # В самом конце работы вставляется дата из названия
                    list_of_final_lists.append(final_list)
                    print("the end of", final_list[0], "||", the_poll['question'])
                    if log_messages:
                        log_messages.insert(tkinter.END, f"{final_list[0]} || {the_poll['question']}\n")
        print("right before the 'infi' if statement")
        if not (infi and stopper_flag):
            break
    return list_of_final_lists

def one_poll_in_excel(poll_results, ws, color_list, log_messages, year=2024):
    light_blue = PatternFill(fill_type="solid", start_color="c6e2ff")
    split_day = poll_results[0].split(".")
    # print(split_day)
    dance_day = datetime.strptime(f"{split_day[0]}/{split_day[1]}/{year}",
                                  "%d/%m/%Y").weekday()  # Day of the week that we have parsed
    #     dance_day = poll_results[0].weekday()
    if dance_day in color_list:
        log_messages.insert(tkinter.END, f"coloring {poll_results[0]}; The date is {dance_day}\n")
        to_color = True
    else:
        to_color = False
    # for cols in ws.iter_cols(min_row=1, max_col=100, max_row=1): #первый цикл,
    # где я ищу самый последний столбик, перед которым создам новый
    if ws.cell(1,
               ws.max_column - 3).value == "Посещено":
        added_cols_letter = ws.cell(1, ws.max_column - 3).column_letter
        ws.insert_cols(ws.max_column - 3)  # Создание таблицы
        ws.cell(1, ws.max_column - 4).number_format = "d-m"
        ws.cell(1, ws.max_column - 4).font = Font(bold=True)
        ws.column_dimensions[added_cols_letter].width = 7  # Установка нужной ширины столбца

        for i in range(1, len(poll_results)):  # ЦИКЛ ДЛЯ КАЖДОГО ЧЕЛОВЕКА + дата в начале столбика
            #             print(i, poll_results[i-1])
            insane_string = f'{added_cols_letter}{i}'
            # offset(columns =- 1) is there cause after inserting i don't have the actual index for filling
            # insane_string это строка, меняющаяся от второго до 29 индекса в пределах одного столбика.
            # Каждая строка описывает одну ячейку

            ws[insane_string] = poll_results[i - 1]  # ВСТАВКА ЗНАЧЕНИЯ

            # Форматирование
            if to_color:  # Проверка на заполнение цветом
                ws[insane_string].fill = light_blue
            ws[insane_string].border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws[insane_string].alignment = Alignment(horizontal='center', vertical='center')  # центрирование
        ws[f"{added_cols_letter}{ws.max_row}"] = f"=SUM({added_cols_letter}2:{added_cols_letter}{ws.max_row - 1})"


def vk_api_creation(link):
    return vk_api.VkApi(token=link[link.find("=") + 1:link.find("&")])

def forming_a_new_list(ids, exception, session_vk):
    chat = session_vk.method("messages.getChat", {"chat_id": ids})
    people = []
    for persons_id in chat['users']:
        if persons_id not in exception:  # Денис Бакунин и Людмилла Одинцова
            peep = session_vk.method("users.get", {"user_ids": persons_id})[0]
            people.append(f"{peep['first_name']} {peep['last_name']}")
    return people

def list_formation(session_vk, log_messages, ids=29, exception=None, did_anything_change=False):
    if exception is None:
        exception = [83919720, 700518702]
    if did_anything_change:
        log_messages.insert(tkinter.END, "Reformed a list of conv. members\n")
        list_of_dancers_array = forming_a_new_list(ids=ids, exception=exception, session_vk=session_vk)
    elif ids == 29:
        log_messages.insert(tkinter.END, "\nUsed the old conv. list of members\n")
        list_of_dancers_array = DANCE_MEMBERS
    elif ids == 70:
        log_messages.insert(tkinter.END, "\nUsed the old conv. list of members\n")
        list_of_dancers_array = FENCE_MEMBERS

    amount = len(list_of_dancers_array)
    dict_of_dancers = dict(zip(list_of_dancers_array, [i for i in range(amount + 1)]))
    return dict_of_dancers
def sheet_action_func(action, people, wb, excel_path_to_save):
    ws_old = wb[wb.sheetnames[len(wb.sheetnames)-1]]
    if action:
        blue = PatternFill(fill_type="solid", start_color="6fa8dc")
        gray = PatternFill(fill_type="solid", start_color="D3D3D3")
        white = PatternFill(fill_type="solid", start_color="ffffff")
        ws = wb.create_sheet("2024")
        list_len = len(people)
        for row in range(1, len(people)+2):
            ws.cell(row, 1).fill = gray
            ws.cell(row, 1).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
            ws.cell(row, 3).fill = blue
            ws.cell(row, 3).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
            ws.cell(row, 4).fill = blue
            ws.cell(row, 4).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
            ws.cell(row, 5).fill = blue
            ws.cell(row, 5).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
            ws.cell(row, 6).fill = blue
            ws.cell(row, 6).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
        ws.cell(1, 1).value = "ФИО"
        ws.cell(1, 1).fill = white
        ws.column_dimensions[ws.cell(1, 1).column_letter].width = 26
        ws.column_dimensions[ws.cell(1, 2).column_letter].width = 7
        ws.cell(1, 3).value = "Посещено"
        ws.cell(1, 4).value = "Пропущено"
        ws.cell(1, 5).value = "Голосование проигнорировано"
        ws.cell(1, 6).value = "В %"

        for row in range(1, len(people)+3):
            ws.cell(row, 2).value = ws_old.cell(row, ws.max_column-4).value
            ws.cell(row, 2).border = Border(top=Side(style='thin', color="000000"),  # Установка сетки
                                              right=Side(style='thin', color="000000"),
                                              left=Side(style='thin', color="000000"),
                                              bottom=Side(style='thin', color="000000"))
            ws.cell(row, 2).alignment = Alignment(horizontal='center', vertical='center')  # центрирование
        ws.cell(1, 2).font = Font(bold=True)
        ws.cell(ws.max_row, 2).border = Border(  # Установка сетки
                                        right=Side(style='thin', color="ffffff"),
                                        left=Side(style='thin', color="ffffff"),
                                        bottom=Side(style='thin', color="ffffff"))
        if type(ws.cell(1, 2).value == datetime):
            ws.cell(1, 2).value = ws_old.cell(1, ws.max_column - 4).value.strftime(
                '%d.%m')  # in case i have a date, not a string

        for row in range(2, list_len+2):
            ws.cell(row, 1).value = list(people)[row-2]
        # wb.save(excel_path_to_save)
        return ws
    else:
        return ws_old

# def one_column_workaround(column, end_excel_index, ws, first_word, color=None):
#     ws.cell(1, column).value = first_word
#     for row in range(2, end_excel_index+1):
#         pass

def insert_all_the_polls(excel_path, path_to_save, link, log_messages=None, did_anything_change=False, chat_option="", sheet_action=False, infi=False, color_list=None):
    if color_list is None:
        color_list = []

    chat_id, exceptions = OPTIONS[chat_option] # getting the needed information for whatever chat i got
    try:
        wb = load_workbook(excel_path)  # Загрузил файл и поместил в переменную "Tantsy.xlsx" "vkApi_excel.xlsx"
        session = vk_api_creation(link=link)
        dict_of_members = list_formation(ids=chat_id % 2000000000, exception=exceptions, session_vk=session, log_messages=log_messages, did_anything_change=did_anything_change)

        ws = sheet_action_func(sheet_action, dict_of_members.keys(), wb, path_to_save) # making a sheet or ignoring and moving on
        # wb = load_workbook(excel_path)
        # ws = wb[wb.sheetnames[len(wb.sheetnames)-1]]
        # ws = wb[wb.sheetnames[sheet_number]]  # Получил доступ к странице
        # ws = wb[wb.sheetnames[len(wb.sheetnames)]]



        log_messages.insert("0.end", f"Working with {ws.title}\n"f"The stopper is gonna be {ws.cell(1, ws.max_column - 4).value}")
        # log_messages.insert(tkinter.END, f"Working with {ws.title}\n")
        # log_messages.insert(tkinter.END, f"The stopper is gonna be {ws.cell(1, ws.max_column - 4).value.strftime('%d.%m')}\n")
        stopper = find_date_re(ws.cell(1, ws.max_column - 4).value)
        list_of_final_lists = getting_history(stopper,
                                              dict_of_dancers=dict_of_members,
                                              infi=infi,
                                              log_messages=log_messages,
                                              session=session,
                                              chat_id=chat_id)

        # ----------------------------------------------------------------------------------
        for i in [-x for x in range(1, len(list_of_final_lists) + 1)]:
            one_poll_in_excel(poll_results=list_of_final_lists[i], ws=ws, year=wb.sheetnames[len(wb.sheetnames)-1][:4],
                              color_list=color_list,
                              log_messages=log_messages)
        # ----------------------------------------------------------------------------------
        for row in range(2, ws.max_row):
            all_data = f'{ws.cell(row, 2).coordinate}:{ws.cell(row, ws.max_column - 4).coordinate}'
            insane_percentage_string = f'=SUM({all_data})/(COUNTA({all_data}))*100'
            ws.cell(row,
                    ws.max_column - 3).value = f"=SUM({ws.cell(row, 2).coordinate}:{ws.cell(row, ws.max_column - 4).coordinate})"
            ws.column_dimensions[ws.cell(1, ws.max_column - 3).column_letter].width = 11
            ws.cell(row,
                    ws.max_column - 2).value = f'=COUNTIF({ws.cell(row, 2).coordinate}:{ws.cell(row, ws.max_column - 4).coordinate},"н")'
            ws.column_dimensions[ws.cell(1, ws.max_column - 2).column_letter].width = 12
            ws.cell(row,
                    ws.max_column - 1).value = f'=COUNTIF({ws.cell(row, 2).coordinate}:{ws.cell(row, ws.max_column - 4).coordinate},"и") '
            ws.column_dimensions[ws.cell(1, ws.max_column - 1).column_letter].width = 30
            # Я задаю 30, потому что 30*7 = 210 пикселей. Почему-то оно вычисляется так и меня это раздражает
            ws.cell(row, ws.max_column).value = insane_percentage_string
        wb.save(path_to_save)
        log_messages.insert(tkinter.END, "Proccess complete")
    except (exceptions.InvalidFileException, vk_api.exceptions.ApiError) as e:
        messagebox.showerror(message=f"{e}")
        log_messages.delete('1.0', tkinter.END) # deletes all the text
        return

